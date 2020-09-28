# 爬蟲用到的library
from urllib.request import urlopen
import re
# 建Excel用到的library
import openpyxl
from openpyxl import  Workbook

# 爬蟲部分
# 使用函式讀取html原始碼，配合regex選取所需內容並返回形態爲dict的資料
def parse_meta(entry):
    return {
        # 選取(<td width="8%" align="center">)與( </td>)之間的內容
        'names': re.findall(r'<td width="8%" align="center">(.*?) </td>', entry),
        # 選取(考區 : )與(</a>)之間的內容
        'testAreas': re.findall(r'考區 : (.*?)</a>', entry),
        # 選取(scope="row"><div align="center" class=)與(</div>)之間的內容
        'selectionResults': re.findall(r'scope="row"><div align="center" class=(.*?)</div>', entry)
    }

# https://www.com.tw/cross/check_046522_NO_1_109_0_3.html
# https://www.com.tw/cross/check_046492_NO_1_108_0_3.html
# https://www.com.tw/cross/check_046482_NO_1_107_0_3.html
# https://www.com.tw/cross/check_046362_NO_1_106_0_3.html
# (046522，046492，046482，046362）：各年度的校系代碼不一而且不具邏輯變化，故無法使用迴圈一次抓取所有年度的資料

# 建立函式方便重複使用
def crawling(url):
    # 讀取url內容並寫入content
    content = urlopen(url).read().decode('utf-8')

    # 呼叫函式建立dict
    meta = parse_meta(content)
    meta_size = len(meta['names'])

    # selectionResults中的內容包含leftred/leftgreen標籤，配合迴圈再次使用regex選取(正x、備x)等內容
    for i in range(meta_size):
        # re.findall會返回list，加上[0]讀取其中的字串並寫入原字串
        meta['selectionResults'][i] = re.findall(r'">(.*)', meta['selectionResults'][i])[0]

    # 預覽爬蟲結果
    for i in range(meta_size):
        print(meta['names'][i], meta['testAreas'][i], meta['selectionResults'][i])
    
    # 返回資料以及其數量
    return meta, meta_size

# ----------------------------

# 取得各年份資料及其數量
meta_109, meta_109_size = crawling("https://www.com.tw/cross/check_046522_NO_1_109_0_3.html")
meta_108, meta_108_size = crawling("https://www.com.tw/cross/check_046492_NO_1_108_0_3.html")
meta_107, meta_107_size = crawling("https://www.com.tw/cross/check_046482_NO_1_107_0_3.html")
meta_106, meta_106_size = crawling("https://www.com.tw/cross/check_046362_NO_1_106_0_3.html")
meta_105, meta_105_size = crawling("https://www.com.tw/cross/check_046332_NO_1_105_0_3.html")
meta_104, meta_104_size = crawling("https://www.com.tw/cross/check_046332_NO_1_104_0_3.html")
meta_103, meta_103_size = crawling("https://www.com.tw/cross/check_046292_NO_1_103_0_3.html")
meta_102, meta_102_size = crawling("https://www.com.tw/cross/check_046212_NO_1_102_0_3.html")

# 輸出檔案部分
# 建立Excel空白活頁簿
wb = Workbook()
# 建立各年份的工作表
sheet = wb.active
sheet_109 = wb.create_sheet("109", 0); sheet_108 = wb.create_sheet("108"); sheet_107 = wb.create_sheet("107")
sheet_106 = wb.create_sheet("106"); sheet_105 = wb.create_sheet("105"); sheet_104 = wb.create_sheet("104")
sheet_103 = wb.create_sheet("103"); sheet_102 = wb.create_sheet("102")
wb.remove(sheet)

# 填入第一列的欄位名稱
sheet_109['A1'] = '考生姓名'; sheet_109['B1'] = '考區'; sheet_109['C1'] = '甄試結果'
sheet_108['A1'] = '考生姓名'; sheet_108['B1'] = '考區'; sheet_108['C1'] = '甄試結果'
sheet_107['A1'] = '考生姓名'; sheet_107['B1'] = '考區'; sheet_107['C1'] = '甄試結果'
sheet_106['A1'] = '考生姓名'; sheet_106['B1'] = '考區'; sheet_106['C1'] = '甄試結果'
sheet_105['A1'] = '考生姓名'; sheet_105['B1'] = '考區'; sheet_105['C1'] = '甄試結果'
sheet_104['A1'] = '考生姓名'; sheet_104['B1'] = '考區'; sheet_104['C1'] = '甄試結果'
sheet_103['A1'] = '考生姓名'; sheet_103['B1'] = '考區'; sheet_103['C1'] = '甄試結果'
sheet_102['A1'] = '考生姓名'; sheet_102['B1'] = '考區'; sheet_102['C1'] = '甄試結果'

# 使用迴圈逐年份逐列填入資料
for i in range(meta_109_size):
    name = meta_109['names'][i]; testArea = meta_109['testAreas'][i]; selectionResult = meta_109['selectionResults'][i]
    sheet_109.append([name, testArea, selectionResult])

for i in range(meta_108_size):
    name = meta_108['names'][i]; testArea = meta_108['testAreas'][i]; selectionResult = meta_108['selectionResults'][i]
    sheet_108.append([name, testArea, selectionResult])

for i in range(meta_107_size):
    name = meta_107['names'][i]; testArea = meta_107['testAreas'][i]; selectionResult = meta_107['selectionResults'][i]
    sheet_107.append([name, testArea, selectionResult])

for i in range(meta_106_size):
    name = meta_106['names'][i]; testArea = meta_106['testAreas'][i]; selectionResult = meta_106['selectionResults'][i]
    sheet_106.append([name, testArea, selectionResult])
for i in range(meta_105_size):
    name = meta_105['names'][i]; testArea = meta_105['testAreas'][i]; selectionResult = meta_105['selectionResults'][i]
    sheet_105.append([name, testArea, selectionResult])
for i in range(meta_104_size):
    name = meta_104['names'][i]; testArea = meta_104['testAreas'][i]; selectionResult = meta_104['selectionResults'][i]
    sheet_104.append([name, testArea, selectionResult])
for i in range(meta_103_size):
    name = meta_103['names'][i]; testArea = meta_103['testAreas'][i]; selectionResult = meta_103['selectionResults'][i]
    sheet_103.append([name, testArea, selectionResult])
for i in range(meta_103_size):
    name = meta_103['names'][i]; testArea = meta_103['testAreas'][i]; selectionResult = meta_103['selectionResults'][i]
    sheet_103.append([name, testArea, selectionResult])
for i in range(meta_102_size):
    name = meta_102['names'][i]; testArea = meta_102['testAreas'][i]; selectionResult = meta_102['selectionResults'][i]
    sheet_102.append([name, testArea, selectionResult])

# 儲存成XLSX檔
wb.save("測資.xlsx")


